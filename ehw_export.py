#!/usr/bin/env python3
import sys
import subprocess
import importlib
import json
import pandas as pd
from pathlib import Path
import shutil
import pwd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import FormulaRule
import re
from ehw_transform import build_yearly_view, build_monthly_view

import ehw_fix_Images

# --- Debug flag ---
DEBUG = bool(os.getenv("EHW_DEBUG"))
DEBUG = DEBUG

# Image management settings
IMG_MODE = os.getenv("EHW_IMG_MODE", "copy").lower()  # copy | symlink
PRUNE = bool(os.getenv("EHW_PRUNE"))

CONFIG_FILE = "./ehw_export.conf.json"
MAX_XLSX_FILES = 10

# === Automatische Paketinstallation ===
def ensure_package(pkg):
    try:
        importlib.import_module(pkg)
    except ImportError:
        print(f"[INFO] Installiere fehlendes Paket: {pkg}")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])

for package in ["pandas", "openpyxl"]:
    ensure_package(package)

# --- Datum & Werte Parsing ---
def parse_date_variants(date_str):
    if not date_str:
        return "", "", "", ""
    try:
        dt = datetime.fromisoformat(date_str.replace("Z", ""))
        if dt.hour == 0 and dt.minute == 0:
            orig = dt.strftime("%d.%m.%Y")
        else:
            orig = dt.strftime("%d.%m.%Y %H:%M")
        return orig, dt.strftime("%Y"), dt.strftime("%Y-%m"), dt.strftime("%Y-%m-%d")
    except Exception:
        return date_str, "", "", ""

def parse_value_numeric(val):
    if val is None:
        return None
    try:
        num_str = re.sub(r"[^0-9,.\-]", "", str(val)).replace(",", ".")
        return float(num_str) if num_str else None
    except ValueError:
        return None

import re as _re

def safe_name(s: str) -> str:
    """Sanitize names for filesystem paths."""
    if s is None:
        return "unknown"
    return _re.sub(r"[\\/|:*?\"<>]", "_", str(s)).strip()

def resolve_room_and_object(counter: dict, room_map: dict) -> tuple[str | None, str | None]:
    """
    Determine the room and object names for a given counter.
    Returns (room_name, object_name).
    """
    room_id = counter.get("roomId")
    room_name = room_map.get(room_id)
    cn = (counter.get("counterName") or "").strip()

    # Helper to extract prefix from a name using . or -
    def extract_prefix(name: str) -> str:
        for delim in ('.', '-'):
            if delim in name:
                return name.split(delim, 1)[0]
        return name

    if room_name:
        object_name = extract_prefix(room_name)
    else:
        object_name = extract_prefix(cn) if cn else None

    return room_name, object_name

# --- Canonical index builder for hidden store ---
import os as _os
import re as _re2

# Build an index of images in the hidden canonical store, supporting multiple layouts:
# 1) flat:   <root>/<obj_uuid>_<room_uuid>_<filename>
# 2) nested: <root>/<obj_uuid>/<room_uuid>/<filename>
# 3) minimal (fallback): <root>/<filename>
CANON_FLAT_RE = _re2.compile(r"^(?P<obj>[0-9a-f\-]{8,})_(?P<room>[0-9a-f\-]{8,})_(?P<file>.+)$", _re2.IGNORECASE)

def build_canonical_index(canonical_root: Path) -> dict:
    idx = {
        "by_tuple": {},      # (obj_uuid, room_uuid, file) -> Path
        "by_room_file": {},  # (room_uuid, file) -> list[Path]
        "by_file": {},       # file -> list[Path]
    }
    if not canonical_root.exists():
        return idx
    for dirpath, _dirs, files in _os.walk(canonical_root):
        dpath = Path(dirpath)
        # Case 2 and 3: nested or minimal
        parts = dpath.relative_to(canonical_root).parts
        obj_uuid = parts[0] if len(parts) >= 1 and parts[0] else None
        room_uuid = parts[1] if len(parts) >= 2 and parts[1] else None
        for fn in files:
            p = dpath / fn
            m = CANON_FLAT_RE.match(fn)
            if m:
                o = m.group("obj"); r = m.group("room"); f = m.group("file")
                idx["by_tuple"][(o, r, f)] = p
                idx["by_room_file"].setdefault((r, f), []).append(p)
                idx["by_file"].setdefault(f, []).append(p)
                continue
            # nested case
            f = fn
            if obj_uuid and room_uuid:
                idx["by_tuple"][(obj_uuid, room_uuid, f)] = p
                idx["by_room_file"].setdefault((room_uuid, f), []).append(p)
            # minimal fallback
            idx["by_file"].setdefault(f, []).append(p)
    return idx

from pathlib import Path as _Path
import shutil as _shutil

def copy_canonical_image(
    local_image_file_name: str | None,
    canonical_root: Path,
    object_uuid: str | None,
    room_uuid: str | None,
    object_name: str | None,
    room_name: str | None,
    counter_name: str | None,
    date_full: str | None,
    visible_root: Path,
    verbose: bool = False,
    canon_idx: dict | None = None
) -> Path | None:
    """Copy or symlink from <target>/.<folder>/<obj>_<room>_<file> to <target>/<folder>/<Object>/<Room>/<file>.
    Returns dest path as Path or None if source doesn't exist or no filename.
    """
    if not local_image_file_name or not object_uuid or not room_uuid:
        return None
    file_name = _Path(local_image_file_name).name
    src = None
    if canon_idx:
        # 1) exact tuple
        src = canon_idx["by_tuple"].get((object_uuid, room_uuid, file_name))
        # 2) by (room_uuid, file)
        if not src:
            cands = canon_idx["by_room_file"].get((room_uuid, file_name)) or []
            src = cands[0] if cands else None
        # 3) by file only (last resort)
        if not src:
            cands = canon_idx["by_file"].get(file_name) or []
            src = cands[0] if cands else None
    if not src:
        # Legacy flat name fallback
        canonical_name = f"{object_uuid}_{room_uuid}_{file_name}"
        src = canonical_root / canonical_name
    if verbose:
        print(f"[IMG] try src={src}")
    if not src or not Path(src).exists():
        if verbose:
            print(f"[IMG] missing: {src}")
        return None
    dest_dir = visible_root / safe_name(object_name) / safe_name(room_name)
    if verbose:
        print(f"[IMG] ensure dest_dir={dest_dir}")
    dest_dir.mkdir(parents=True, exist_ok=True)
    safe_counter = safe_name(counter_name) if counter_name else "unknown"
    safe_date = date_full.replace("-", "") if date_full else "nodate"
    new_name = f"{safe_counter}_{safe_date}_{file_name}"
    dest = dest_dir / new_name
    if not dest.exists():
        if IMG_MODE == "symlink":
            if verbose:
                print(f"[IMG] symlink {src} -> {dest}")
            dest.symlink_to(src)
        else:
            if verbose:
                print(f"[IMG] copy {src} -> {dest}")
            _shutil.copy2(src, dest)
    else:
        if verbose:
            print(f"[IMG] exists: {dest}")
    return dest
# --- Helpers ---
def load_config(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def cleanup_old_excels(folder: Path, prefix: str):
    files = sorted(folder.glob(f"{prefix}*.xlsx"), key=os.path.getmtime)
    while len(files) > MAX_XLSX_FILES:
        old_file = files.pop(0)
        try:
            old_file.unlink()
            print(f"[Cleanup] removed {old_file.name}")
        except Exception as e:
            print(f"[WARN] Konnte {old_file} nicht löschen: {e}")

# --- Hauptlogik ---
ALL_ROWS = []

def process_folder(sync_dir: Path, target_base_dir: Path):
    json_path = sync_dir / f"{sync_dir.name}.json"
    print("\n----------------------------------------------")
    print(f"Bearbeite Ordner: {sync_dir.name}")
    print("----------------------------------------------")
    if not json_path.exists():
        print(f"[WARN] JSON fehlt: {json_path}")
        return

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    # oldest & newest entry date
    all_dates = []
    for c in data.get("counters", []):
        for e in c.get("entries", {}).get("entries", []):
            _o,_y,_ym,df = parse_date_variants(e.get("date"))
            if df:
                all_dates.append(df)
    if all_dates:
        print(f"  Ältester Wert: {min(all_dates)}")
        print(f"  Neuester Wert: {max(all_dates)}")

    # Map each roomId (UUID) to its cleartext name
    room_map = {}
    for room in data.get("rooms", []):
        rid = room.get("roomId")
        rname = room.get("name") or room.get("title")
        if rid:
            room_map[rid] = rname
    # canonical source and visible target roots for images
    object_uuid = data.get("objectId")
    canonical_root = target_base_dir / f".{object_uuid}"
    visible_images_root = target_base_dir
    if DEBUG:
        print(f"[DBG] canonical_root={canonical_root}")
        print(f"[DBG] visible_root={visible_images_root}")
        print(f"[DBG] object_uuid={object_uuid}")
    canon_idx = build_canonical_index(canonical_root)
    if DEBUG:
        print(f"[DBG] canonical indexed: tuples={len(canon_idx['by_tuple'])} by_room={len(canon_idx['by_room_file'])} by_file={len(canon_idx['by_file'])}")

    expected_files: list[Path] = []
    rows = []
    for counter in data.get("counters", []):
        for entry in counter.get("entries", {}).get("entries", []):
            date_orig, date_year, date_yearmonth, date_full = parse_date_variants(entry.get("date"))
            value_orig = entry.get("value")
            value_num = parse_value_numeric(value_orig)
            # Resolve human-readable room and object names
            room_name, object_name = resolve_room_and_object(counter, room_map)
            dest_path: Path | None = copy_canonical_image(
                entry.get("localImageFileName"),
                canonical_root=canonical_root,
                object_uuid=object_uuid,
                room_uuid=counter.get("roomId"),
                object_name=object_name,
                room_name=room_name,
                counter_name=counter.get("counterName"),
                date_full=date_full,
                visible_root=visible_images_root,
                verbose=DEBUG,
                canon_idx=canon_idx,
            )
            if dest_path:
                expected_files.append(dest_path)
                try:
                    rel = dest_path.relative_to(target_base_dir)
                except Exception:
                    rel = dest_path
                img_link_formula = f'=HYPERLINK("{rel.as_posix()}","Bild")'
            else:
                img_link_formula = ""

            # Build row dictionary (CounterUUID suppressed)
            row = {
                "Object": object_name,
                "Room": room_name,
                "CounterName": counter.get("counterName"),
                "Bild": "Bild" if dest_path else "",  # text; hyperlink applied later
                "CounterType": counter.get("counterType"),
                "CounterUnit": counter.get("counterUnit"),
                "CounterId": counter.get("counterId"),
                "RoomId": counter.get("roomId"),
                "Date_Orig": date_orig,
                "Date_Year": date_year,
                "Date_YearMonth": date_yearmonth,
                "Date_Full": date_full,
                "Value_Orig": value_orig,
                "Value_Num": value_num,
            }
            # store absolute path for post-write hyperlinking
            if dest_path:
                row["__BildAbs"] = dest_path.resolve().as_posix()
            rows.append(row)

    # --- Virtual counter processing ---
    # Build lookup of counters by UUID
    counter_by_uuid = {c.get("uuid"): c for c in data.get("counters", [])}

    # Extract virtual counters
    virtual_counters = [c for c in data.get("counters", []) if c.get("counterType") == "VIRTUAL"]

    # Helper: extract entries as dict date_full -> numeric value
    def extract_value_map(counter_obj):
        m = {}
        for e in counter_obj.get("entries", {}).get("entries", []):
            _o, _y, _ym, dfull = parse_date_variants(e.get("date"))
            v = parse_value_numeric(e.get("value"))
            if dfull and v is not None:
                m[dfull] = v
        return m

    # Process each virtual counter
    for vctr in virtual_counters:
        vcdata = vctr.get("virtualCounterData") or {}
        master_uuid = vcdata.get("masterCounterUuid")
        add_uuids = vcdata.get("counterUuidsToBeAdded") or []
        sub_uuids = vcdata.get("counterUuidsToBeSubtracted") or []

        if not master_uuid or master_uuid not in counter_by_uuid:
            continue  # skip invalid

        master = counter_by_uuid.get(master_uuid)
        adds = [counter_by_uuid[u] for u in add_uuids if u in counter_by_uuid]
        subs = [counter_by_uuid[u] for u in sub_uuids if u in counter_by_uuid]

        # --- Unit fallback: derive from master name ---
        unit = master.get("counterUnit")
        if not unit:
            mname = master.get("counterName", "").lower()
            if "wasser" in mname:
                unit = "m3"
            elif "strom" in mname:
                unit = "kwh"
            elif "wärme" in mname or "waerme" in mname:
                unit = "kwh"
            else:
                unit = "unknown"

        # Build value maps
        mv = extract_value_map(master)
        add_maps = [extract_value_map(c) for c in adds]
        sub_maps = [extract_value_map(c) for c in subs]

        # Collect all dates
        all_dates = sorted(set(mv.keys()) | set().union(*[m.keys() for m in add_maps]) | set().union(*[m.keys() for m in sub_maps]))

        # Room/Object resolution
        room_name, object_name = resolve_room_and_object(vctr, room_map)
        # fallback if room not found
        if not room_name:
            room_name = vctr.get("counterName") or "Virtual"
        if not object_name and room_name:
            object_name = room_name.split('.', 1)[0]

        for d in all_dates:
            base_val = mv.get(d)
            if base_val is None:
                continue

            add_val = sum(am.get(d, 0) for am in add_maps)
            sub_val = sum(sm.get(d, 0) for sm in sub_maps)
            virt_val = base_val + add_val - sub_val

            # Build row
            row = {
                "Object": object_name,
                "Room": room_name,
                "CounterName": vctr.get("counterName"),
                "Bild": "",
                "CounterType": "VIRTUAL",
                "CounterUnit": unit,
                "CounterId": "",
                "RoomId": vctr.get("roomId"),
                "Date_Orig": datetime.strptime(d, "%Y-%m-%d").strftime("%d.%m.%Y"),
                "Date_Year": d[:4],
                "Date_YearMonth": d[:7],
                "Date_Full": d,
                "Value_Orig": virt_val,
                "Value_Num": virt_val,
            }
            rows.append(row)

    df = pd.DataFrame(rows)
    # Sort the raw dataframe for cleaner Excel output
    if "Room" in df.columns and "CounterName" in df.columns and "Date_Full" in df.columns:
        df = df.sort_values(["Room", "CounterName", "Date_Full"], ascending=[True, True, True])
    # --- Add delta/prev/days for raw data ---
    from ehw_transform import add_delta_columns
    df = add_delta_columns(df)
    df["_SourceFolder"] = sync_dir.name
    ALL_ROWS.append(df.copy())
    print("  --- Zählerübersicht ---")
    grouped = {}
    for idx, r in df.iterrows():
        key = (r["Room"], r["CounterName"])
        grouped.setdefault(key, {"count":0, "last":None})
        grouped[key]["count"] += 1
        grouped[key]["last"] = r["Date_Full"]
    for (room, cname), info in sorted(grouped.items(), key=lambda x: (str(x[0][0]), str(x[0][1]))):
        last = str(info["last"]).split(" ")[0] if info["last"] else "-"
        print(f"    {room}  T:{cname}  #{info['count']}  last:{last}")
    # Ensure column order and keep helper path for now
    desired_order = [
        "Object", "Room", "CounterName", "Bild",
        "CounterType", "CounterUnit", "CounterId", "RoomId",
        "Date_Orig", "Date_Year", "Date_YearMonth", "Date_Full",
        "Value_Orig", "Value_Num", "__BildAbs"
    ]
    df = df.reindex(columns=[c for c in desired_order if c in df.columns])

    # Optional prune of stale images in visible tree
    if PRUNE and visible_images_root.exists():
        keep = {p.resolve() for p in expected_files}
        removed = 0
        for dirpath, _dirs, files in _os.walk(visible_images_root):
            dpath = Path(dirpath)
            for fn in files:
                p = (dpath / fn).resolve()
                if p not in keep:
                    if DEBUG:
                        print(f"[PRUNE] remove stale: {p}")
                    try:
                        p.unlink()
                        removed += 1
                    except Exception as e:
                        print(f"[WARN] prune failed for {p}: {e}")
        if DEBUG:
            print(f"[PRUNE] removed {removed} files")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_name = f"##{sync_dir.name}-{timestamp}.xlsx"
    excel_path = target_base_dir / excel_name

    export_with_format(df, excel_path, script_name="ehw_export")
    cleanup_old_excels(target_base_dir, f"##{sync_dir.name}-")

    # Also create/overwrite a stable latest file for cross-links
    latest_path = target_base_dir / f"{sync_dir.name}.xlsx"
    try:
        shutil.copy2(excel_path, latest_path)
        print(f"[OK] latest XLS updated: {latest_path.name}")
    except Exception as e:
        print(f"[WARN] Konnte Latest Excel nicht schreiben: {latest_path} -> {e}")

# --- Excel-Formatierung ---
def export_with_format(df, file_path, script_name):
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=2, sheet_name="Zählerdaten")
        # --- Build yearly and monthly aggregated counter sheets ---
        try:
            df_year = build_yearly_view(df)
            df_year.to_excel(writer, index=False, sheet_name="Zählerdaten_Jahr")
        except Exception as e:
            print(f"[WARN] Jahres-Ansicht konnte nicht erzeugt werden: {e}")

        try:
            df_month = build_monthly_view(df)
            df_month.to_excel(writer, index=False, sheet_name="Zählerdaten_Monat")
        except Exception as e:
            print(f"[WARN] Monats-Ansicht konnte nicht erzeugt werden: {e}")


        # After write: turn Bild text into hyperlinks using __BildAbs
        ws = writer.book.active
        # Find column indices
        headers = [cell.value for cell in ws[3]]
        def col_index(name):
            return headers.index(name) + 1 if name in headers else None
        col_bild = col_index("Bild")
        col_abs = col_index("__BildAbs")
        if col_bild and col_abs:
            for r in range(4, ws.max_row + 1):
                text = ws.cell(row=r, column=col_bild).value
                path = ws.cell(row=r, column=col_abs).value
                if text and path:
                    # Use absolute file URI to avoid locale/comma/semicolon formula issues
                    uri = f"file://{path}"
                    ws.cell(row=r, column=col_bild).value = text
                    ws.cell(row=r, column=col_bild).hyperlink = uri
        # Drop helper column if present by clearing its header
        if col_abs:
            ws.delete_cols(col_abs, 1)

    wb = load_workbook(file_path)
    ws = wb.active
    # --- Re-add header info in row 1 ---
    try:
        header_info = f"{script_name} -- {file_path.parent} -- {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} -- {pwd.getpwuid(os.getuid()).pw_name} -- {os.uname().nodename}"
        ws["A1"] = header_info
    except Exception as e:
        print(f"[WARN] Header row could not be written: {e}")
    # --- Minimal Excel Table Only ---
    from openpyxl.worksheet.table import Table, TableStyleInfo

    # Determine table range using non-empty header columns
    header_cells = ws[3]
    actual_cols = 0
    for idx, cell in enumerate(header_cells, start=1):
        if cell.value not in (None, "", "__BildAbs"):
            actual_cols = idx

    if actual_cols > 0 and ws.max_row >= 3:
        last_row = ws.max_row
        last_col_letter = ws.cell(row=3, column=actual_cols).column_letter
        # Force full range so virtual rows always included
        table_range = f"A3:{last_col_letter}{last_row}"

        # Always use fixed table name "tblEHW" for Zählerdaten
        tbl = Table(displayName="tblEHW", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tbl.tableStyleInfo = style
        ws.add_table(tbl)

    # --- Add tables for yearly and monthly sheets ---
    from openpyxl.utils import get_column_letter

    def add_table_to_sheet(sheet_name, table_style):
        if sheet_name not in wb.sheetnames:
            return
        wsx = wb[sheet_name]
        if wsx.max_row < 2 or wsx.max_column < 1:
            return

        # Determine last column/row
        last_col_letter = get_column_letter(wsx.max_column)
        last_row = wsx.max_row
        table_range = f"A1:{last_col_letter}{last_row}"

        # Assign fixed table names
        if sheet_name == "Zählerdaten_Jahr":
            table_name = "tblehwJahr"
        elif sheet_name == "Zählerdaten_Monat":
            table_name = "tblehwMonat"
        elif sheet_name == "Zählerdaten":
            table_name = "tblEHW"
        else:
            table_name = "TblData"

        t = Table(displayName=table_name, ref=table_range)
        t.tableStyleInfo = TableStyleInfo(
            name=table_style,
            showRowStripes=True,
            showColumnStripes=False
        )
        wsx.add_table(t)

    # Different colors for the tables
    add_table_to_sheet("Zählerdaten_Jahr", "TableStyleMedium4")
    add_table_to_sheet("Zählerdaten_Monat", "TableStyleMedium7")

    wb.save(file_path)
    print(f"[OK] XLS saved: {file_path.name}")

# --- Main ---
def get_version():
    """Return version based on git tag or fallback VERSION file."""
    root = Path(__file__).parent

    # First try: git describe
    try:
        import subprocess
        version = subprocess.check_output(
            ["git", "describe", "--tags", "--dirty", "--always"],
            cwd=root,
            stderr=subprocess.DEVNULL,
        ).decode("utf-8").strip()
        if version:
            return version
    except Exception:
        pass

    # Second try: VERSION file
    vfile = root / "VERSION"
    try:
        return vfile.read_text(encoding="utf-8").strip()
    except:
        return "unknown"

def main():
    config = load_config(path=CONFIG_FILE)
    version = get_version()
    print("==============================================")
    print(f"EHW Exporter v{version}")
    print("Konfigurationsdatei:", CONFIG_FILE)
    print("==============================================")
    base_sync_dir = Path(config["source_base_dir"]).resolve()
    target_base_dir = Path(config["target_base_dir"]).resolve()
    target_base_dir.mkdir(parents=True, exist_ok=True)
    if DEBUG:
        print("[DBG] EHW_DEBUG=1 (image copy debug enabled)")
    subprocess.run(
        ["python3", "ehw_fix_Images.py", "-c", CONFIG_FILE, "--apply"],
        check=True
    )
    for folder_name in config["folders"]:
        sync_dir = base_sync_dir / folder_name
        if sync_dir.exists():
            process_folder(sync_dir, target_base_dir)
        else:
            print(f"[WARN] Ordner fehlt: {sync_dir}")

    if ALL_ROWS:
        combined = pd.concat(ALL_ROWS, ignore_index=True)

        # Remove _SourceFolder entirely
        if "_SourceFolder" in combined.columns:
            combined = combined.drop(columns=["_SourceFolder"])

        combined_path = target_base_dir / "ehw+.xlsx"
        export_with_format(combined, combined_path, script_name="ehw_export_combined")
        print(f"[OK] combined XLS saved: {combined_path.name}")
if __name__ == "__main__":
    main()

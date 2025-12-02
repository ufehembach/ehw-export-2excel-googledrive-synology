def build_virtual_mapping(counters_json):
    """
    Returns two dicts:
      virtual_to_physical: {virtual_uuid: [phys_uuid1, phys_uuid2, ...]}
      physical_to_virtual: {phys_uuid: virtual_uuid}
    """
    virtual_to_physical = {}
    physical_to_virtual = {}

    for c in counters_json:
        cid = c.get("uuid")
        vdata = c.get("virtualCounterData") or {}
        add_list = vdata.get("counterUuidsToBeAdded") or []
        sub_list = vdata.get("counterUuidsToBeSubtracted") or []
        children = list(add_list) + list(sub_list)

        if children:
            virtual_to_physical[cid] = children
            for child in children:
                physical_to_virtual[child] = cid

    return virtual_to_physical, physical_to_virtual

import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import re as _re

# Helper function for virtual reset detection
def detect_virtual_reset(counter_id, previous_val, current_val, vmap):
    """
    Detect reset for virtual counters:
    If counter is virtual AND current_val < previous_val → reset.
    """
    if counter_id in vmap:
        if previous_val is not None and current_val is not None:
            if current_val < previous_val:
                return True
    return False

def build_yearly_view(df):
    df2 = df.copy()
    df2["Date"] = pd.to_datetime(df2["Date_Full"], errors="coerce")
    df2 = df2.sort_values(["CounterId", "Date"])
    df2["Year"] = df2["Date"].dt.year

    yearly = []
    for cid, group in df2.groupby("CounterId"):
        for year, g2 in group.groupby("Year"):
            year_end = pd.Timestamp(year=year, month=12, day=31)
            g_before = g2[g2["Date"] <= year_end]

            if not g_before.empty:
                row = g_before.iloc[-1]
            else:
                g_after = group[
                    (group["Date"] > year_end) &
                    (group["Date"] <= year_end + pd.Timedelta(days=15))
                ]
                if g_after.empty:
                    continue
                row = g_after.iloc[0]

            yearly.append({
                "Object": row["Object"],
                "Room": row["Room"],
                "CounterName": row["CounterName"],
                "CounterId": cid,
                "Year": year,
                "Date": row["Date"],
                "Value_Num": row["Value_Num"],
                "CounterType": row["CounterType"],
                "CounterUnit": row["CounterUnit"],
            })

    df_year = pd.DataFrame(yearly).sort_values(["CounterId", "Year"])
    # --- Add delta/prev with reset detection and ResetDetected ---
    prev_vals = []
    prev_dates = []
    deltas = []
    reset_flags = []

    last_value = {}
    last_date = {}

    try:
        vmap = virtual_to_physical
    except:
        vmap = {}

    for idx, row in df_year.iterrows():
        key = row["CounterId"]
        current_val = row["Value_Num"]
        current_date = row["Date"]

        if key not in last_value:
            prev_vals.append(None)
            prev_dates.append(None)
            deltas.append(None)
            reset_flags.append(False)
            last_value[key] = current_val
            last_date[key] = current_date
            continue

        previous_val = last_value[key]
        previous_date = last_date[key]

        # virtual reset detection
        virt_reset = detect_virtual_reset(key, previous_val, current_val, vmap)

        if current_val < previous_val or virt_reset:
            delta = current_val
            prev_vals.append(None)
            prev_dates.append(None)
            reset_flag = True
        else:
            delta = current_val - previous_val
            prev_vals.append(previous_val)
            prev_dates.append(previous_date)
            reset_flag = False

        deltas.append(delta)
        reset_flags.append(reset_flag)

        last_value[key] = current_val
        last_date[key] = current_date

    df_year["PrevValue"] = prev_vals
    df_year["PrevDate"] = prev_dates
    df_year["Delta"] = deltas
    df_year["ResetDetected"] = reset_flags

    # --- Add DeltaPerDay ---
    delta_per_day = []
    for idx, row in df_year.iterrows():
        if row["PrevDate"] is None or row["Delta"] is None:
            delta_per_day.append(None)
        else:
            days = (row["Date"] - row["PrevDate"]).days
            if days > 0:
                delta_per_day.append(row["Delta"] / days)
            else:
                delta_per_day.append(None)
    df_year["DeltaPerDay"] = delta_per_day

    # --- Add Days ---
    days_list = []
    for idx, row in df_year.iterrows():
        if row["PrevDate"] is None:
            days_list.append(None)
        else:
            days_list.append((row["Date"] - row["PrevDate"]).days)
    df_year["Days"] = days_list

    # --- Add Bemerkung and CreatedAt columns ---
    df_year["Bemerkung"] = ""
    df_year["CreatedAt"] = pd.Timestamp.now().isoformat()

    # --- qbm/m3 unit division and Bemerkung update ---
    mask = df_year["CounterUnit"].astype(str).str.lower().isin(["qbm", "m3", "m³", "m^3"])
    if mask.any():
        for col in ["Value_Num", "PrevValue", "Delta"]:
            if col in df_year.columns:
                df_year.loc[mask, col] = df_year.loc[mask, col] / 1000
        df_year.loc[mask, "Bemerkung"] = df_year.loc[mask, "Bemerkung"].astype(str) + "Wasser geteilt durch 1000; "

    # --- Column ordering ---
    desired_order = [
        "Object", "Room", "CounterName", "CounterId",
        "CounterType", "CounterUnit", "Year", "Date", "Value_Num",
        "PrevValue", "PrevDate", "Delta", "DeltaPerDay", "Days",
        "ResetDetected",
        "Bemerkung", "CreatedAt",
    ]
    cols = [c for c in desired_order if c in df_year.columns] + \
           [c for c in df_year.columns if c not in desired_order]

    df_year = df_year[cols]
    # --- Sort with virtual counters first ---
    df_year["_virt"] = df_year["CounterType"].astype(str).str.upper().eq("VIRTUAL").astype(int)
    df_year = df_year.sort_values(["_virt", "CounterName"])
    df_year = df_year.drop(columns=["_virt"])
    # --- Hierarchische Sortierung ---
    # Für virtuelle Zähler und ihre physikalischen Unterzähler eine SortKey erzeugen
    df_year["_SortKey"] = ""

    # Mapping früher gebaut: virtual_to_physical
    try:
        vmap = virtual_to_physical
    except:
        vmap = {}

    order_index = 1
    for cname in sorted(df_year["CounterName"].unique()):
        subset = df_year[df_year["CounterName"] == cname]
        cid = subset["CounterId"].iloc[0]

        if cid in vmap:  # virtual
            base = f"{order_index}.0."
            df_year.loc[df_year["CounterId"] == cid, "_SortKey"] = base + "0"
            # Kinder einsortieren
            sub_idx = 1
            for child in vmap[cid]:
                df_year.loc[df_year["CounterId"] == child, "_SortKey"] = base + str(sub_idx)
                sub_idx += 1
            order_index += 1
        else:
            base = f"{order_index}.0."
            df_year.loc[df_year["CounterId"] == cid, "_SortKey"] = base + "0"
            order_index += 1

    df_year = df_year.sort_values("_SortKey")
    df_year = df_year.drop(columns=["_SortKey"])
    return df_year


def build_monthly_view(df):
    df2 = df.copy()
    df2["Date"] = pd.to_datetime(df2["Date_Full"], errors="coerce")
    df2 = df2.sort_values(["CounterId", "Date"])
    df2["YearMonth"] = df2["Date"].dt.to_period("M")

    monthly = []
    for cid, group in df2.groupby("CounterId"):
        min_month = group["YearMonth"].min()
        max_month = group["YearMonth"].max()

        for ym in pd.period_range(min_month, max_month, freq="M"):
            month_end = ym.to_timestamp(how="end")
            g_before = group[group["Date"] <= month_end]

            if not g_before.empty:
                row = g_before.iloc[-1]
            else:
                g_after = group[
                    (group["Date"] > month_end) &
                    (group["Date"] <= month_end + pd.Timedelta(days=10))
                ]
                if g_after.empty:
                    continue
                row = g_after.iloc[0]

            monthly.append({
                "Object": row["Object"],
                "Room": row["Room"],
                "CounterName": row["CounterName"],
                "CounterId": cid,
                "YearMonth": str(ym),
                "Date": row["Date"],
                "Value_Num": row["Value_Num"],
                "CounterType": row["CounterType"],
                "CounterUnit": row["CounterUnit"],
            })

    df_month = pd.DataFrame(monthly).sort_values(["CounterId", "YearMonth"])

    # --- Add delta/prev with reset detection and ResetDetected ---
    prev_vals = []
    prev_dates = []
    deltas = []
    reset_flags = []

    last_value = {}
    last_date = {}

    try:
        vmap = virtual_to_physical
    except:
        vmap = {}

    for idx, row in df_month.iterrows():
        key = row["CounterId"]
        current_val = row["Value_Num"]
        current_date = row["Date"]

        if key not in last_value:
            # first entry → no delta
            prev_vals.append(None)
            prev_dates.append(None)
            deltas.append(None)
            reset_flags.append(False)
            last_value[key] = current_val
            last_date[key] = current_date
            continue

        previous_val = last_value[key]
        previous_date = last_date[key]

        virt_reset = detect_virtual_reset(key, previous_val, current_val, vmap)

        if current_val < previous_val or virt_reset:
            # reset detected
            delta = current_val
            prev_vals.append(None)
            prev_dates.append(None)
            reset_flag = True
        else:
            delta = current_val - previous_val
            prev_vals.append(previous_val)
            prev_dates.append(previous_date)
            reset_flag = False

        deltas.append(delta)
        reset_flags.append(reset_flag)

        last_value[key] = current_val
        last_date[key] = current_date

    df_month["PrevValue"] = prev_vals
    df_month["PrevDate"] = prev_dates
    df_month["Delta"] = deltas
    df_month["ResetDetected"] = reset_flags

    # --- Add DeltaPerDay (Verbrauch pro Tag) ---
    delta_per_day = []
    for idx, row in df_month.iterrows():
        if row["PrevDate"] is None or row["Delta"] is None:
            delta_per_day.append(None)
        else:
            days = (row["Date"] - row["PrevDate"]).days
            if days > 0:
                delta_per_day.append(row["Delta"] / days)
            else:
                delta_per_day.append(None)
    df_month["DeltaPerDay"] = delta_per_day

    # --- Add Days (Anzahl Tage zwischen Ablesungen) ---
    days_list = []
    for idx, row in df_month.iterrows():
        if row["PrevDate"] is None:
            days_list.append(None)
        else:
            days_list.append((row["Date"] - row["PrevDate"]).days)
    df_month["Days"] = days_list

    # --- Add Bemerkung and CreatedAt columns ---
    df_month["Bemerkung"] = ""
    df_month["CreatedAt"] = pd.Timestamp.now().isoformat()

    # --- qbm/m3 unit division and Bemerkung update ---
    mask = df_month["CounterUnit"].astype(str).str.lower().isin(["qbm", "m3", "m³", "m^3"])
    if mask.any():
        for col in ["Value_Num", "PrevValue", "Delta"]:
            if col in df_month.columns:
                df_month.loc[mask, col] = df_month.loc[mask, col] / 1000
        df_month.loc[mask, "Bemerkung"] = df_month.loc[mask, "Bemerkung"].astype(str) + "Wasser geteilt durch 1000; "

    # --- Column ordering ---
    desired_order = [
        "Object", "Room", "CounterName", "CounterId",
        "CounterType", "CounterUnit", "YearMonth", "Date", "Value_Num",
        "PrevValue", "PrevDate", "Delta", "DeltaPerDay", "Days",
        "ResetDetected",
        "Bemerkung", "CreatedAt",
    ]
    cols = [c for c in desired_order if c in df_month.columns] + \
           [c for c in df_month.columns if c not in desired_order]

    df_month = df_month[cols]
    # --- Sort with virtual counters first ---
    df_month["_virt"] = df_month["CounterType"].astype(str).str.upper().eq("VIRTUAL").astype(int)
    df_month = df_month.sort_values(["_virt", "CounterName"])
    df_month = df_month.drop(columns=["_virt"])
    # --- Hierarchische Sortierung ---
    # Für virtuelle Zähler und ihre physikalischen Unterzähler eine SortKey erzeugen
    df_month["_SortKey"] = ""

    # Mapping früher gebaut: virtual_to_physical
    try:
        vmap = virtual_to_physical
    except:
        vmap = {}

    order_index = 1
    for cname in sorted(df_month["CounterName"].unique()):
        subset = df_month[df_month["CounterName"] == cname]
        cid = subset["CounterId"].iloc[0]

        if cid in vmap:  # virtual
            base = f"{order_index}.0."
            df_month.loc[df_month["CounterId"] == cid, "_SortKey"] = base + "0"
            # Kinder einsortieren
            sub_idx = 1
            for child in vmap[cid]:
                df_month.loc[df_month["CounterId"] == child, "_SortKey"] = base + str(sub_idx)
                sub_idx += 1
            order_index += 1
        else:
            base = f"{order_index}.0."
            df_month.loc[df_month["CounterId"] == cid, "_SortKey"] = base + "0"
            order_index += 1

    df_month = df_month.sort_values("_SortKey")
    df_month = df_month.drop(columns=["_SortKey"])
    return df_month

def extract_unit(counter_name: str) -> str:
    """
    Extracts the unit (Wohnungseinheit), e.g. DBMP.EG or H1.Whg2
    from CounterName like 'DBMP.EG.Wasser-Küche'.
    """
    if not isinstance(counter_name, str):
        return ""
    parts = counter_name.split(".")
    if len(parts) >= 2:
        return ".".join(parts[:2])
    return parts[0]


def extract_art(counter_type: str, counter_name: str) -> str:
    """
    Detects water/heating/electricity from CounterType or CounterName.
    """
    text = f"{counter_type} {counter_name}".lower()
    if "wasser" in text or "water" in text:
        return "wasser"
    if "wärme" in text or "waerme" in text or "heat" in text:
        return "wärme"
    if "strom" in text or "electric" in text:
        return "strom"
    return ""

def build_summary_table(df_month):
    """
    Builds a clean pivoted summary:
    YearMonth | wasser | wärme | strom | Einheit
    """
    df2 = df_month.copy()
    df2["Ablesung"] = df2["Date"].dt.strftime("%Y.%m")
    df2["Einheit"] = df2["CounterName"].apply(extract_unit)
    df2["Art"] = df2.apply(lambda r: extract_art(r["CounterType"], r["CounterName"]), axis=1)

    summary = (
        df2.pivot_table(
            index="Ablesung",
            columns="Art",
            values="Value_Num",
            aggfunc="max"
        )
        .reset_index()
        .rename_axis(None, axis=1)
    )

    einheit_map = df2.groupby("Ablesung")["Einheit"].first()
    summary["Einheit"] = summary["Ablesung"].map(einheit_map)

    summary = summary.sort_values("Ablesung")
    return summary

def add_delta_columns(df):
    """
    Adds PrevValue, PrevDate, Delta, DeltaPerDay, Days to raw df.
    Reset detection included.
    """
    df2 = df.copy()
    df2["Date_Full"] = pd.to_datetime(df2["Date_Full"], errors="coerce")
    df2 = df2.sort_values(["CounterId", "Date_Full"])

    prev_vals = []
    prev_dates = []
    deltas = []
    reset_flags = []
    last_value = {}
    last_date = {}

    for idx, row in df2.iterrows():
        key = row["CounterId"]
        current_val = row.get("Value_Num", None)
        current_date = row["Date_Full"]

        if key not in last_value:
            prev_vals.append(None)
            prev_dates.append(None)
            deltas.append(None)
            reset_flags.append(False)
            last_value[key] = current_val
            last_date[key] = current_date
            continue

        previous_val = last_value[key]
        previous_date = last_date[key]

        if current_val is None or previous_val is None:
            prev_vals.append(None)
            prev_dates.append(None)
            deltas.append(None)
            reset_flags.append(False)
        elif current_val < previous_val:
            reset_flag = True
            prev_vals.append(None)
            prev_dates.append(None)
            deltas.append(current_val)
            reset_flags.append(reset_flag)
        else:
            reset_flag = False
            prev_vals.append(previous_val)
            prev_dates.append(previous_date)
            deltas.append(current_val - previous_val)
            reset_flags.append(reset_flag)

        last_value[key] = current_val
        last_date[key] = current_date

    df2["PrevValue"] = prev_vals
    df2["PrevDate"] = prev_dates
    df2["Delta"] = deltas
    df2["ResetDetected"] = reset_flags

    # --- qbm/m3 unit division and Bemerkung update ---
    mask = df2["CounterUnit"].astype(str).str.lower().isin(["qbm", "m3", "m³", "m^3"])
    if mask.any():
        for col in ["Value_Num", "PrevValue", "Delta"]:
            if col in df2.columns:
                df2.loc[mask, col] = df2.loc[mask, col] / 1000
        df2.loc[mask, "Bemerkung"] = df2.loc[mask, "Bemerkung"].astype(str) + "Wasser geteilt durch 1000; "

    delta_per_day = []
    days_list = []
    for idx, row in df2.iterrows():
        if row["PrevDate"] is None or row["Delta"] is None:
            delta_per_day.append(None)
            days_list.append(None)
        else:
            days = (row["Date_Full"] - row["PrevDate"]).days
            days_list.append(days if days > 0 else None)
            if days and days > 0:
                delta_per_day.append(row["Delta"] / days)
            else:
                delta_per_day.append(None)

    df2["DeltaPerDay"] = delta_per_day
    df2["Days"] = days_list

    # --- Add Bemerkung and CreatedAt columns ---
    df2["Bemerkung"] = ""
    df2["CreatedAt"] = pd.Timestamp.now().isoformat()

    # --- Column ordering ---
    desired_order = [
        "Object", "Room", "CounterName", "CounterId",
        "CounterType", "CounterUnit", "Date_Full", "Value_Num",
        "PrevValue", "PrevDate", "Delta", "DeltaPerDay", "Days",
        "ResetDetected",
        "Bemerkung", "CreatedAt",
    ]
    cols = [c for c in desired_order if c in df2.columns] + \
           [c for c in df2.columns if c not in desired_order]

    df2 = df2[cols]
    # --- Sort with virtual counters first ---
    df2["_virt"] = df2["CounterType"].astype(str).str.upper().eq("VIRTUAL").astype(int)
    df2 = df2.sort_values(["_virt", "CounterName"])
    df2 = df2.drop(columns=["_virt"])
    # --- Hierarchische Sortierung ---
    # Für virtuelle Zähler und ihre physikalischen Unterzähler eine SortKey erzeugen
    df2["_SortKey"] = ""

    # Mapping früher gebaut: virtual_to_physical
    try:
        vmap = virtual_to_physical
    except:
        vmap = {}

    order_index = 1
    for cname in sorted(df2["CounterName"].unique()):
        subset = df2[df2["CounterName"] == cname]
        cid = subset["CounterId"].iloc[0]

        if cid in vmap:  # virtual
            base = f"{order_index}.0."
            df2.loc[df2["CounterId"] == cid, "_SortKey"] = base + "0"
            # Kinder einsortieren
            sub_idx = 1
            for child in vmap[cid]:
                df2.loc[df2["CounterId"] == child, "_SortKey"] = base + str(sub_idx)
                sub_idx += 1
            order_index += 1
        else:
            base = f"{order_index}.0."
            df2.loc[df2["CounterId"] == cid, "_SortKey"] = base + "0"
            order_index += 1

    df2 = df2.sort_values("_SortKey")
    df2 = df2.drop(columns=["_SortKey"])
    return df2

def add_table_to_sheet(sheet_name, table_style):
    if sheet_name not in wb.sheetnames:
        return
    wsx = wb[sheet_name]
    if wsx.max_row < 2 or wsx.max_column < 1:
        return

    last_col_letter = get_column_letter(wsx.max_column)
    last_row = wsx.max_row
    table_range = f"A1:{last_col_letter}{last_row}"

    if sheet_name == "Zählerdaten_Jahr":
        table_name = "tblehwJahr"
    elif sheet_name == "Zählerdaten_Monat":
        table_name = "tblehwMonat"
    else:
        table_name = "TblData"

    t = Table(displayName=table_name, ref=table_range)
    t.tableStyleInfo = TableStyleInfo(
        name=table_style,
        showRowStripes=True,
        showColumnStripes=False
    )
    wsx.add_table(t)
